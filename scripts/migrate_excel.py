#!/usr/bin/env python3
"""
Парсинг Excel-эталонов в PostgreSQL с embeddings.
Учитывает структуру: Класс → Атрибуты → Required → Пример
"""

import sys
import json
import argparse
from pathlib import Path
from typing import List, Dict, Any, Optional

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("Установите openpyxl: pip install openpyxl")
    sys.exit(1)

# Добавляем путь к core
sys.path.insert(0, str(Path(__file__).parent.parent / 'scripts'))
from core.vector_store import VectorStore


class EtalonParser:
    """Парсер Excel-файла с эталонами."""
    
    def __init__(self, excel_path: Path):
        self.excel_path = excel_path
        self.wb = load_workbook(excel_path)
        self.ws = self.wb.active
        
    def parse_all(self) -> List[Dict]:
        """Парсить все классы и примеры из Excel."""
        etalons = []
        current_class = None
        current_attributes = []
        current_required = []
        
        rows = list(self.ws.iter_rows(values_only=True))
        i = 0
        
        while i < len(rows):
            row = rows[i]
            a_val = row[0] if row[0] else None
            
            # Найти начало нового класса
            if a_val and isinstance(a_val, str) and a_val.strip() and not a_val.startswith('МТР'):
                # Это название класса
                current_class = a_val.strip()
                
                # Следующая строка - заголовки атрибутов
                if i + 1 < len(rows):
                    attr_row = rows[i + 1]
                    current_attributes = []
                    for j, val in enumerate(attr_row):
                        if val and isinstance(val, str) and val.strip():
                            current_attributes.append(val.strip())
                
                # Строка через одну - required flags
                if i + 2 < len(rows):
                    req_row = rows[i + 2]
                    current_required = []
                    for j, val in enumerate(req_row):
                        if j < len(current_attributes):
                            req = val and str(val).strip().lower() in ['да', 'true', 'yes']
                            current_required.append(req)
                
                # Строка через две - пример
                if i + 3 < len(rows):
                    example_row = rows[i + 3]
                    example_name = example_row[0] if example_row[0] else ""
                    
                    if example_name and isinstance(example_name, str):
                        # Собрать атрибуты
                        attributes = {}
                        for j, attr_name in enumerate(current_attributes):
                            if j + 3 < len(example_row) and example_row[j + 3]:
                                attributes[attr_name] = str(example_row[j + 3])
                        
                        etalons.append({
                            'class_name': current_class,
                            'example_name': example_name,
                            'unit': str(example_row[2]) if len(example_row) > 2 and example_row[2] else "",
                            'status': str(example_row[3]) if len(example_row) > 3 and example_row[3] else "",
                            'attributes': attributes,
                            'attribute_specs': {
                                attr: {'required': req}
                                for attr, req in zip(current_attributes, current_required)
                            },
                            'excel_row': i + 1
                        })
                
                # Переместиться к следующему классу (пропустить 4 строки)
                i += 5
            else:
                i += 1
        
        return etalons
    
    def get_stats(self) -> Dict:
        """Получить статистику по файлу."""
        etalons = self.parse_all()
        classes = set(e['class_name'] for e in etalons)
        
        return {
            'total_examples': len(etalons),
            'unique_classes': len(classes),
            'total_attributes': sum(len(e['attributes']) for e in etalons),
            'avg_attributes_per_example': sum(len(e['attributes']) for e in etalons) / len(etalons) if etalons else 0
        }


def generate_embedding(text: str, provider: str = "ollama") -> List[float]:
    """Генерация embedding для текста."""
    if provider == "ollama":
        import requests
        try:
            response = requests.post(
                "http://localhost:11434/api/embeddings",
                json={"model": "qwen3-embedding:4b", "prompt": text},
                timeout=30
            )
            response.raise_for_status()
            embedding = response.json().get('embedding', [])
            # Обрезать до 2000d (максимум для HNSW в pgvector)
            if len(embedding) > 2000:
                embedding = embedding[:2000]
            return embedding
        except Exception:
            pass
    
    # Fallback: случайный вектор для тестирования
    import random
    random.seed(text)
    return [random.uniform(-1, 1) for _ in range(2000)]


def migrate_to_postgresql(etalons: List[Dict], vector_store: VectorStore):
    """Миграция эталонов в PostgreSQL с embeddings."""
    print(f"Начинаю миграцию {len(etalons)} эталонов...")
    
    for i, etalon in enumerate(etalons, 1):
        if i % 100 == 0:
            print(f"  Обработано: {i}/{len(etalons)}")
        
        # Генерация embedding
        embedding = generate_embedding(etalon['example_name'])
        
        # Сохранение в PostgreSQL
        vector_store.add(
            code=f"etalon_{i:04d}",
            name=etalon['example_name'],
            class_name=etalon['class_name'],
            attributes=etalon['attributes'],
            embedding=embedding
        )
    
    print(f"✅ Миграция завершена. Всего записей: {vector_store.count()}")


def main():
    parser = argparse.ArgumentParser(description='Миграция Excel-эталонов в PostgreSQL')
    parser.add_argument('--excel', type=Path, required=True, help='Путь к Excel-файлу')
    parser.add_argument('--db-host', default='localhost', help='Хост PostgreSQL')
    parser.add_argument('--db-name', default='nomenclature', help='Имя базы данных')
    parser.add_argument('--db-user', default='postgres', help='Пользователь')
    parser.add_argument('--db-password', default='', help='Пароль')
    parser.add_argument('--stats-only', action='store_true', help='Только статистика')
    
    args = parser.parse_args()
    
    # Парсинг
    print(f"Парсинг {args.excel}...")
    parser = EtalonParser(args.excel)
    
    if args.stats_only:
        stats = parser.get_stats()
        print("\nСтатистика:")
        for key, val in stats.items():
            print(f"  {key}: {val}")
        return
    
    etalons = parser.parse_all()
    print(f"Найдено эталонов: {len(etalons)}")
    
    # Подключение к PostgreSQL
    vector_store = VectorStore(
        host=args.db_host,
        database=args.db_name,
        user=args.db_user,
        password=args.db_password
    )
    
    # Инициализация схемы
    print("Инициализация схемы PostgreSQL...")
    vector_store.init_schema()
    
    # Миграция
    migrate_to_postgresql(etalons, vector_store)
    
    # Итоговая статистика
    stats = vector_store.get_stats()
    print("\nИтоговая статистика:")
    for key, val in stats.items():
        print(f"  {key}: {val}")


if __name__ == '__main__':
    main()